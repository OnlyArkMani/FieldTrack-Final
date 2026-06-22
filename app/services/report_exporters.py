"""Render a ReportData to CSV / Excel / PDF bytes.

All three are SYNCHRONOUS and CPU-bound; callers run them via
asyncio.to_thread (see run_report_job) so the event loop never stalls.

PALETTE (FieldTrack brand — kept in sync with the apps' ThemeData):
  amber  #F5A623  header fill / accents
  cream  #FFF8E7  zebra-stripe fill
  dark   #1A1A2E  body text
  purple #8B7FD4  summary accent
"""
from __future__ import annotations

import csv
import io
from typing import Any

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from app.schemas.report import ReportFormat
from app.services.report_service import ReportData, ReportTable

# Brand colours
AMBER = "F5A623"
CREAM = "FFF8E7"
DARK = "1A1A2E"
PURPLE = "8B7FD4"
WHITE = "FFFFFF"
# Compliance conditional-formatting tints (light backgrounds, readable text).
LIGHT_GREEN = "E6F4EA"  # visited
LIGHT_CORAL = "FBE3E1"  # not visited


def render_report(data: ReportData, fmt: ReportFormat) -> bytes:
    if fmt is ReportFormat.CSV:
        return _render_csv(data)
    if fmt is ReportFormat.EXCEL:
        return _render_excel(data)
    return _render_pdf(data)


# ── CSV ──────────────────────────────────────────────────────────────────
def _render_csv(data: ReportData) -> bytes:
    """One file, sectioned: header → summary → each table. utf-8-sig so Excel
    opens it with correct encoding."""
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([f"FieldTrack — {data.title}"])
    w.writerow([data.subtitle])
    if data.filters_text:
        w.writerow(["Filters: " + "  |  ".join(data.filters_text)])
    w.writerow(["Generated", data.generated_at.strftime("%Y-%m-%d %H:%M UTC")])
    w.writerow([])

    if data.summary:
        w.writerow(["Summary"])
        for label, value in data.summary:
            w.writerow([label, value])
        w.writerow([])

    for table in data.tables:
        w.writerow([table.name])
        w.writerow(table.columns)
        for row in table.rows:
            w.writerow(["" if c is None else _csv_safe(c) for c in row])
        w.writerow([])

    return buf.getvalue().encode("utf-8-sig")


# Cell values starting with these characters are interpreted as formulas by
# Excel/Sheets/LibreOffice when a CSV/XLSX is opened (CSV/formula injection).
# Free-text user input (e.g. work_summary) flows into these exports, so any
# string cell starting with one of these is neutralized with a leading
# apostrophe, which Excel renders literally and treats as "text".
_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def _csv_safe(value: Any) -> Any:
    if isinstance(value, str) and value.startswith(_FORMULA_PREFIXES):
        return "'" + value
    return value


# ── Excel ────────────────────────────────────────────────────────────────
def _render_excel(data: ReportData) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor=AMBER)
    header_font = Font(bold=True, color=WHITE, size=11)
    zebra_fill = PatternFill("solid", fgColor=CREAM)
    title_font = Font(bold=True, color=DARK, size=14)
    label_font = Font(bold=True, color=DARK)
    # Conditional-formatting fills for the compliance report (row_styles).
    visited_fill = PatternFill("solid", fgColor=LIGHT_GREEN)
    not_visited_fill = PatternFill("solid", fgColor=LIGHT_CORAL)
    summary_fill = PatternFill("solid", fgColor=AMBER)
    summary_font = Font(bold=True, color=DARK)

    wb = Workbook()

    # ── Summary sheet (first) ──
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = f"FieldTrack — {data.title}"
    ws["A1"].font = title_font
    ws["A2"] = data.subtitle
    ws["A3"] = "  |  ".join(data.filters_text)
    ws["A4"] = "Generated " + data.generated_at.strftime("%Y-%m-%d %H:%M UTC")
    r = 6
    for label, value in data.summary:
        ws.cell(row=r, column=1, value=label).font = label_font
        ws.cell(row=r, column=2, value=value)
        r += 1
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 40

    # ── One sheet per table ──
    for table in data.tables:
        sheet = wb.create_sheet(title=_safe_sheet_name(table.name))
        sheet.append(table.columns)
        for col_idx in range(1, len(table.columns) + 1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        n_cols = len(table.columns)
        for i, row in enumerate(table.rows, start=2):
            sheet.append(["" if c is None else _csv_safe(c) for c in row])
            # Conditional formatting (compliance) overrides zebra striping when a
            # per-row style tag is present; otherwise fall back to zebra stripes.
            style_tag = (
                table.row_styles[i - 2]
                if table.row_styles and (i - 2) < len(table.row_styles)
                else None
            )
            if style_tag == "summary":
                for col_idx in range(1, n_cols + 1):
                    cell = sheet.cell(row=i, column=col_idx)
                    cell.fill = summary_fill
                    cell.font = summary_font
            elif style_tag == "visited":
                for col_idx in range(1, n_cols + 1):
                    sheet.cell(row=i, column=col_idx).fill = visited_fill
            elif style_tag == "not_visited":
                for col_idx in range(1, n_cols + 1):
                    sheet.cell(row=i, column=col_idx).fill = not_visited_fill
            elif i % 2 == 0:  # zebra stripe even data rows (default tables)
                for col_idx in range(1, n_cols + 1):
                    sheet.cell(row=i, column=col_idx).fill = zebra_fill
            for c in table.numeric_cols:
                sheet.cell(row=i, column=c + 1).alignment = Alignment(
                    horizontal="right"
                )

        # Auto column width from content (capped so a long work-summary doesn't
        # blow the layout out).
        for col_idx in range(1, len(table.columns) + 1):
            longest = len(str(table.columns[col_idx - 1]))
            for row in table.rows:
                if col_idx - 1 < len(row):
                    longest = max(longest, len(str(row[col_idx - 1] or "")))
            sheet.column_dimensions[get_column_letter(col_idx)].width = min(
                max(longest + 2, 10), 50
            )
        sheet.freeze_panes = "A2"  # freeze the header row

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _safe_sheet_name(name: str) -> str:
    # Excel sheet names: <=31 chars, no  : \ / ? * [ ]
    cleaned = "".join(c for c in name if c not in r':\/?*[]')
    return cleaned[:31] or "Sheet"


# ── PDF ──────────────────────────────────────────────────────────────────
def _render_pdf(data: ReportData) -> bytes:
    bio = io.BytesIO()
    doc = SimpleDocTemplate(
        bio,
        pagesize=landscape(A4),
        topMargin=18 * mm,
        bottomMargin=16 * mm,
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        title=data.title,
        author="FieldTrack",
    )

    styles = getSampleStyleSheet()
    brand = ParagraphStyle(
        "Brand", parent=styles["Title"], fontSize=20, textColor=colors.HexColor(f"#{AMBER}"),
        spaceAfter=2,
    )
    h_title = ParagraphStyle(
        "RTitle", parent=styles["Heading2"], textColor=colors.HexColor(f"#{DARK}"),
        spaceAfter=2,
    )
    meta = ParagraphStyle(
        "Meta", parent=styles["Normal"], fontSize=9,
        textColor=colors.HexColor("#6B6B80"),
    )
    section = ParagraphStyle(
        "Section", parent=styles["Heading3"], textColor=colors.HexColor(f"#{PURPLE}"),
        spaceBefore=8, spaceAfter=4,
    )
    cell = ParagraphStyle("Cell", parent=styles["Normal"], fontSize=8, leading=10)

    elements: list = [
        Paragraph("FieldTrack", brand),
        Paragraph(data.title, h_title),
        Paragraph(data.subtitle, meta),
    ]
    if data.filters_text:
        elements.append(Paragraph("Filters: " + "  •  ".join(data.filters_text), meta))
    elements.append(Spacer(1, 6 * mm))

    # Summary statistics box (two-column label/value grid).
    if data.summary:
        elements.append(Paragraph("Summary", section))
        elements.append(_summary_table(data.summary))
        elements.append(Spacer(1, 6 * mm))

    # Data tables.
    for table in data.tables:
        elements.append(Paragraph(table.name, section))
        elements.append(_data_table(table, cell))
        elements.append(Spacer(1, 4 * mm))

    doc.build(elements, onFirstPage=_footer, onLaterPages=_footer)
    return bio.getvalue()


def _summary_table(summary: list[tuple[str, str]]) -> Table:
    # Lay stat pairs out 2-per-row to use the landscape width.
    flat: list = []
    row: list = []
    for label, value in summary:
        row.append(Paragraph(f"<b>{_esc(label)}</b>", _SMALL))
        row.append(Paragraph(_esc(value), _SMALL))
        if len(row) == 4:
            flat.append(row)
            row = []
    if row:
        while len(row) < 4:
            row.append("")
        flat.append(row)

    t = Table(flat, colWidths=[45 * mm, 50 * mm, 45 * mm, 50 * mm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(f"#{CREAM}")),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor(f"#{AMBER}")),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E0DCC8")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
    ]))
    return t


def _data_table(table: ReportTable, cell_style: ParagraphStyle) -> Table:
    header = [Paragraph(f"<b>{_esc(c)}</b>", _HEADER) for c in table.columns]
    body = [header]
    for row in table.rows:
        body.append([
            Paragraph(_esc("" if c is None else str(c)), cell_style)
            for c in row
        ])

    t = Table(body, repeatRows=1)
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{AMBER}")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D8D4C4")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
    ]
    # Zebra striping on the body.
    for i in range(1, len(body)):
        if i % 2 == 0:
            style.append(
                ("BACKGROUND", (0, i), (-1, i), colors.HexColor(f"#{CREAM}"))
            )
    # Right-align numeric columns.
    for c in table.numeric_cols:
        style.append(("ALIGN", (c, 1), (c, -1), "RIGHT"))
    t.setStyle(TableStyle(style))
    return t


def _footer(canvas, doc) -> None:
    """Page-number + brand footer on every page."""
    canvas.saveState()
    canvas.setFont("Helvetica", 8)
    canvas.setFillColor(colors.HexColor("#6B6B80"))
    width, _ = landscape(A4)
    canvas.drawString(12 * mm, 8 * mm, "FieldTrack — Employee Tracking")
    canvas.drawRightString(width - 12 * mm, 8 * mm, f"Page {doc.page}")
    canvas.restoreState()


def _esc(text: str) -> str:
    return (
        str(text)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )


_styles = getSampleStyleSheet()
_SMALL = ParagraphStyle("SmallCell", parent=_styles["Normal"], fontSize=8, leading=10)
_HEADER = ParagraphStyle(
    "HeaderCell", parent=_styles["Normal"], fontSize=8, leading=10,
    textColor=colors.white,
)
